from gui import *
from openpyxl import load_workbook
import iapws
import os

hodnoty = {'p': None, 't1': None, 't2': None, 'Q': None, 'Q_lat': None, 'P': None,
        'h1': None, 'h2': None, 'h3': None, 'v1': None, 'v2': None, 'v3': None,'s1': None, 's2': None, 's3': None}

def compute_button_signal(self):
    self.compute_button.clicked.connect(self.compute)

def save_button_signal(self):
    self.save_button.clicked.connect(self.save_data)

def vypocet_hodnot(teplota_1,tlak,hmot_tok):

    h1 = iapws.iapws97._Region1(teplota_1,tlak)['h']
    h2 = iapws.iapws97._Region4(tlak,0)['h']
    h3 = iapws.iapws97._Region4(tlak,1)['h']

    v1 = iapws.iapws97._Region1(teplota_1,tlak)['v']
    v2 = iapws.iapws97._Region4(tlak,0)['v']
    v3 = iapws.iapws97._Region4(tlak,1)['v']

    s1 = iapws.iapws97._Region1(teplota_1,tlak)['s']
    s2 = iapws.iapws97._Region4(tlak,0)['s']
    s3 = iapws.iapws97._Region4(tlak,1)['s']

    Q_latent = h3-h2
    Q = (h2-h1)+(h3-h2)
    P = Q * hmot_tok
    return Q,Q_latent,P,h1,h2,h3,v1,v2,v3,s1,s2,s3
    #      0,1,       2 3  4  5  6  7  8  9  10 11

def compute(self):
    default_t1 = self.teplota_1_edit.text()
    default_p = self.tlak_edit.text()
    default_hm = self.hmot_edit.text()

    teplota_kr = 647.15 #K
    tlak_kr = 22.064 #MPa

    teplota_tr_bod = 273.15 #K
    tlak_tr_bod = 0.000611 #MPa

    try:
        t1 = float(default_t1)
        p = float(default_p)
        hm = float(default_hm)

    except ValueError:
        self.error_text.append('Vložte číselnou hodnotu')

    else:
        t1 = (float(default_t1)) + 273.15  #na Kelviny
        p = (float(default_p)) /1000  #na MPa

        if t1 > teplota_kr:
            self.error_text.append('Zvolte menší teplotu než 374 °C')
        elif t1 < teplota_tr_bod:
            self.error_text.append('Zvolte větší teplotu než 0 °C')
        elif p > tlak_kr:
            self.error_text.append('Zvolte menší tlak než 22.064 MPa')
        elif p < tlak_tr_bod:
            self.error_text.append('Zvolte větší tlak než 611 Pa')
        else:
            px = (iapws.iapws97._PSat_T(t1))  #Mpa
            tx = (iapws.iapws97._TSat_P(p))   #K

            if px > p:
                lim_p = round((min(px,tlak_kr)) * 1000,3)
                lim_t = round((min(tx,teplota_kr)) - 273.15,3)
                hlaska = ('Jste v oblasti suché páry. \n Zvište tlak nad {}kPa nebo snižte teplotu pod {}°C'.format(lim_p,lim_t))
                self.error_text.append(hlaska)
            else:
                Q = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[0]),3))
                Q_lat = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[1]),3))
                P = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[2]),3))

                h1 = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[3]),3))
                h2 = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[4]),3))
                h3 = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[5]),3))

                v1 = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[6]),6))
                v2 = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[7]),6))
                v3 = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[8]),6))

                s1 = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[9]),6))
                s2 = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[10]),6))
                s3 = str(round((vypocet_hodnot(teplota_1=t1,tlak=p,hmot_tok = hm)[11]),6))

                t3 = str(round((iapws.iapws97._Region4(p,1)['T'])-273.15,3))
                t1_save = t1 - 273.15
                p_save = p * 1000

                vyp_hodnoty = {'p': p_save, 't1': t1_save, 't3': t3, 'Q': Q, 'Q_lat': Q_lat, 'P': P,
                'h1': h1, 'h2': h2, 'h3': h3, 'v1': v1, 'v2': v2, 'v3': v3,'s1': s1, 's2': s2, 's3': s3}
                hodnoty.update(vyp_hodnoty)

                self.celk_teplo_lab.setText(hodnoty.get("Q"))
                self.lat_teplo_lab.setText(hodnoty.get("Q_lat"))
                self.teplota_t2_lab.setText(hodnoty.get("t3"))
                self.vykon_lab.setText(hodnoty.get("P"))


def save_data(self):
    options = QtWidgets.QFileDialog.Options()
    options |= QtWidgets.QFileDialog.DontUseNativeDialog
    fileName, ignored= QtWidgets.QFileDialog.getOpenFileName(None,"QFileDialog.getOpenFileName()", '',"xlsx(*.xlsx)", options=options)

    if fileName == '':
        self.error_text.append('Nevybrán žádný soubor')
    elif self.celk_teplo_lab.text() == '':
        self.error_text.append('Žádná data k uložení')
    else:
        data = load_workbook(fileName)
        self.ws = data.active

        Q = (self.celk_teplo_lab.text())
        Q_lat = (self.lat_teplo_lab.text())
        t3 = (self.teplota_t2_lab.text())
        P = (self.vykon_lab.text())

        t1 = hodnoty.get("t1")
        p = hodnoty.get("p")

        self.ws['B2'] = float(t1)
        self.ws['B3'] = float(p)

        self.ws['B6'] = float(t3)
        self.ws['B7'] = float(Q_lat)
        self.ws['B8'] = float(Q)
        self.ws['B9'] = float(P)

        self.ws['B12'] = float(hodnoty.get("h1"))
        self.ws['B13'] = float(hodnoty.get("h2"))
        self.ws['B14'] = float(hodnoty.get("h3"))

        self.ws['B15'] = float(hodnoty.get("v1"))
        self.ws['B16'] = float(hodnoty.get("v2"))
        self.ws['B17'] = float(hodnoty.get("v3"))

        self.ws['B18'] = float(hodnoty.get("s1"))
        self.ws['B19'] = float(hodnoty.get("s2"))
        self.ws['B20'] = float(hodnoty.get("s3"))

        data.save(fileName)
        self.error_text.append('Data uložena')



Ui_mainWindow.compute_button_signal = compute_button_signal
Ui_mainWindow.compute = compute

Ui_mainWindow.save_button_signal = save_button_signal
Ui_mainWindow.save_data = save_data

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_mainWindow()
    ui.setupUi(MainWindow)
    ui.compute_button_signal()
    ui.save_button_signal()
    MainWindow.show()
    sys.exit(app.exec_())
